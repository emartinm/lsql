# -*- coding: utf-8 -*-
"""
Tests for rankings
"""
from datetime import datetime
import tempfile
import openpyxl

from django.test import TestCase, Client
from django.urls import reverse

from judge.models import NumSolvedCollectionAchievementDefinition, PodiumAchievementDefinition, \
    NumSolvedAchievementDefinition, AchievementDefinition, ObtainedAchievement, Submission, \
    NumSolvedTypeAchievementDefinition, NumSubmissionsProblemsAchievementDefinition
from judge.types import VeredictCode, ProblemType
from judge.tests.test_views import create_user, create_superuser, create_group, create_collection, create_select_problem
from judge.views import first_day_of_course


def create_an_achievement_of_each(coll):
    """Create an achievement of each type: NumSolvedAchievementDefinition, PodiumAchievementDefinition and
    NumSolvedCollectionAchievementDefinition"""
    ach_podium = PodiumAchievementDefinition(name={"es":'Presidente del podio'},
                                             description={"es":'Consigue entrar al podio'},
                                             num_problems=1, position=3)
    ach_collection = NumSolvedCollectionAchievementDefinition(name={"es":'Coleccionista'},
                                                              description={"es":'Resuelve 1 problema\
                                                              de esta coleccion'},
                                                              num_problems=1, collection=coll)
    ach_solved = NumSolvedAchievementDefinition(name={"es":'Resolvista'},
                                                description={"es":'Resuelve 1 problema'},
                                                num_problems=1)
    ach_type = NumSolvedTypeAchievementDefinition(name={"es":'Tipos'},
                                                  description={"es":'Resuelve un problema SELECT'},
                                                  num_problems=1, problem_type=ProblemType.SELECT.name)
    ach_submi_pro = NumSubmissionsProblemsAchievementDefinition(name={"es":'Primer envio'},
                                                                description={"es":'Realiza un envio a \
                                                                cualquier problema'},
                                                                num_submissions=1, num_problems=1)
    ach_type.save()
    ach_submi_pro.save()
    ach_podium.save()
    ach_collection.save()
    ach_solved.save()


class RankingTest(TestCase):
    """Tests for the rankings and achievements"""

    def test_get_podium_achievement(self):
        """Test if get correctly a podium achievement"""
        client = Client()
        podium_achievement = PodiumAchievementDefinition(name={"es":"Top 1"},
                                                         description={"es":'Se el primero'},
                                                         num_problems=1,
                                                         position=1)
        podium_achievement.save()
        coll = create_collection('Test Podium')
        problem = create_select_problem(coll, 'Select 1')
        user = create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        achievements_url = reverse('judge:achievements', args=[user.pk])
        response = client.get(achievements_url, follow=True)
        self.assertIn('Fecha', response.content.decode('utf-8'))

    def test_get_solved_achievement(self):
        """Test if get correctly a solved achievement and check_user function"""
        client = Client()
        solved_achievement_1 = NumSolvedAchievementDefinition(name={"es":'Solo 1'},
                                                             description={"es":'Acierta 1'},
                                                             num_problems=1)
        solved_achievement_2 = NumSolvedAchievementDefinition(name={"es":'Solo 2'},
                                                             description={"es":'Acierta 2'},
                                                             num_problems=2)
        solved_achievement_1.save()
        solved_achievement_2.save()
        coll = create_collection('Test Solved')
        problem = create_select_problem(coll, 'Select 1')
        user = create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        achievements_url = reverse('judge:achievements', args=[user.pk])
        response = client.get(achievements_url, follow=True)
        self.assertIn('Fecha', response.content.decode('utf-8'))
        self.assertIn('Logros pendientes', response.content.decode('utf-8'))

    def test_get_collection_achievement(self):
        """Test if get correctly a collection achievement"""
        client = Client()
        coll = create_collection('Test Solved')
        coll_achievement = NumSolvedCollectionAchievementDefinition(name={"es":'Solo 1'},
                                                                    description={"es":'Acierta 1'},
                                                                    num_problems=1, collection=coll)
        coll_achievement.save()
        problem = create_select_problem(coll, 'Select 1')
        user = create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        achievements_url = reverse('judge:achievements', args=[user.pk])
        response = client.get(achievements_url, follow=True)
        self.assertIn('Fecha', response.content.decode('utf-8'))

    def test_ranking_achievements(self):
        """Test if we can see the number of achievements that have an user at ranking"""
        client = Client()
        user = create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        coll = create_collection('Coleccion de cartas')
        problem = create_select_problem(coll, 'Problema')
        create_an_achievement_of_each(coll)
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        ranking_url = reverse('judge:result', args=[coll.pk])
        group_a = create_group('1A')
        group_a.user_set.add(user)
        response = client.get(ranking_url, follow=True)
        self.assertIn('x5', response.content.decode('utf-8'))

    def test_signal_new_achievement(self):
        """Test for check signals"""
        client = Client()
        create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        coll = create_collection('Coleccion de cartas')
        problem = create_select_problem(coll, 'Problema')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        create_an_achievement_of_each(coll)
        self.assertEqual(ObtainedAchievement.objects.all().count(), 5)

    def test_signal_update_achievement(self):
        """Test for check signals"""
        # Create two users for test all the achievements. Two for the podium
        client = Client()
        user_michu = create_user('passwordmichu', 'michu')
        create_user('passwordimmobile', 'immobile')
        client.login(username='immobile', password='passwordimmobile')
        # Create the Collection for the achievement NumSolvedCollectionAchievementDefinition and Problem
        coll = create_collection('Coleccion de cartas')
        # Create PodiumAchievementDefinition
        ach_podium = PodiumAchievementDefinition(name={"es":'Presidente del podio'},
                                                 description={"es":'Consigue ser el primero'},
                                                 num_problems=1, position=1)
        ach_podium.save()
        # Create NumSolvedCollectionAchievementDefinition
        ach_collection = NumSolvedCollectionAchievementDefinition(name={"es":'Coleccionista'},
                                                                  description={"es":'Resuelve 50\
                                                                  problemas de esta coleccion'},
                                                                  num_problems=50,
                                                                  collection=coll)
        ach_collection.save()
        # Create NumSolvedAchievementDefinition
        ach_solved = NumSolvedAchievementDefinition(name={"es":'Resolvista'},
                                                    description={"es":'Resuelve 50 problemas'},
                                                    num_problems=50)
        ach_solved.save()
        # Create NumSolvedTypeAchievementDefinition
        ach_type = NumSolvedTypeAchievementDefinition(name={"es":'Procedista'},
                                                      description={"es":'Resuelve un problema PROC'},
                                                      num_problems=1, problem_type=ProblemType.PROC.name)
        ach_type.save()
        # Create NumSubmissionsProblemsAchievementDefinition
        ach_submi_pro = NumSubmissionsProblemsAchievementDefinition(name={"es":'Muchos envios'},
                                                                    description={"es":'Envia muchas soluciones'},
                                                                    num_submissions=80, num_problems=1)
        ach_submi_pro.save()
        # Create problem and submit correct answer with "immobile" user, for make this the first to solve the problem
        problem = create_select_problem(coll, 'Problema')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        client.logout()
        # Login with "michu" and submit correct answer. All the checks will be with this user
        client.login(username='michu', password='passwordmichu')
        client.post(submit_select_url, {'code': problem.solution}, follow=True)
        # Whit this definitions our user "michu" don't have any achievement
        self.assertEqual(ObtainedAchievement.objects.filter(user=user_michu).count(), 0)
        # PodiumAchievementDefinition now only need to stay in podium
        # In this test our user "michu" stay at second position, that is why before he didn't have the achievement
        ach_podium.position = 3
        ach_podium.save()
        # NumSolvedCollectionAchievementDefinition only needs one correct submission
        # In this test our user only have one correct submission, that is why before he didn't have the achievement
        ach_collection.num_problems = 1
        ach_collection.save()
        # NumSolvedAchievementDefinition only needs one correct submission
        # In this test our user only have one correct submission, that is why before he didn't have the achievement
        ach_solved.num_problems = 1
        ach_solved.save()
        # NumSolvedTypeAchievementDefinition change to type SELECT
        # In this test our user only resolved a SELECT type problem, not PROC.
        ach_type.problem_type = ProblemType.SELECT.name
        ach_type.save()
        # NumSubmissionsProblemsAchievementDefinition only needs one submission now
        ach_submi_pro.num_submissions = 1
        ach_submi_pro.save()
        # Now our user "michu" have 5 achievements
        self.assertEqual(ObtainedAchievement.objects.filter(user=user_michu).count(), 5)

    def test_not_implemented_raise_achievements(self):
        """Test if check_and_save of AchievementDefinition raise a NotImplementedError"""
        client = Client()
        user = create_user('passwordmichu', 'michu')
        client.login(username='michu', password='passwordmichu')
        achievement_definition = AchievementDefinition(name={"es":'nombre'}, description={"es":'descripcion'})
        with self.assertRaises(NotImplementedError):
            achievement_definition.check_and_save(user)

    def test_obtained_achievements_date(self):
        """Test if the dates of the obtained achievements are correct"""
        user = create_user('passwordmichu', 'michu')
        coll = create_collection('Coleccion de cartas')
        problem_1 = create_select_problem(coll, 'Problema 1')
        problem_2 = create_select_problem(coll, 'Problema 2')
        sub_1 = Submission(code='nada', veredict_code=VeredictCode.AC, user=user, problem=problem_1)
        sub_2 = Submission(code='nada', veredict_code=VeredictCode.AC, user=user, problem=problem_2)
        sub_1.save()
        sub_2.save()
        Submission.objects.filter(id=sub_1.id).update(creation_date=datetime(2006, 3, 5))
        # sub_1_u = submission 1 updated
        sub_1_u = Submission.objects.get(id=sub_1.id)
        Submission.objects.filter(id=sub_2.id).update(creation_date=datetime(2020, 3, 5))
        sub_2_u = Submission.objects.get(id=sub_2.id)

        # Test NumSolvedAchievementDefinition
        ach_solved = NumSolvedAchievementDefinition(name={"es":'Resolvista'},
                                                    description={"es":'Resuelve 2 problemas'},
                                                    num_problems=2)
        ach_solved.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_2_u.creation_date)
        ach_solved.num_problems = 1
        ach_solved.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_1_u.creation_date)
        ObtainedAchievement.objects.all().delete()

        # Test NumSolvedCollectionAchievementDefinition
        ach_coll = NumSolvedCollectionAchievementDefinition(name={"es":'Coleccionista'},
                                                            description={"es":'Resuelve 2 \
                                                            problemas de esta coleccion'},
                                                            num_problems=2, collection=coll)
        ach_coll.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_2_u.creation_date)
        ach_coll.num_problems = 1
        ach_coll.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_1_u.creation_date)
        ObtainedAchievement.objects.all().delete()

        # Test PodiumAchievementDefinition
        ach_podium = PodiumAchievementDefinition(name={"es":'Presidente del podio'},
                                                 description={"es":'Consigue ser el primero'},
                                                 num_problems=2, position=1)
        ach_podium.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_2_u.creation_date)
        ach_podium.num_problems = 1
        ach_podium.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_1_u.creation_date)
        ObtainedAchievement.objects.all().delete()

        # Test NumSolvedTypeAchievementDefinition
        ach_type = NumSolvedTypeAchievementDefinition(name={"es":'Tipos'},
                                                      description={"es":'Resuelve un problema SELECT'},
                                                      num_problems=2, problem_type=ProblemType.SELECT.name)
        ach_type.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_2_u.creation_date)
        ach_type.num_problems = 1
        ach_type.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_1_u.creation_date)
        ObtainedAchievement.objects.all().delete()

        # Test NumSubmissionsProblemsAchievementDefinition
        ach_submissions = NumSubmissionsProblemsAchievementDefinition(name={"es":'Ya van dos problemas'},
                                                  description={"es":'Realiza un envio en dos problemas'},
                                                  num_problems=2, num_submissions=2)
        ach_submissions.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_2_u.creation_date)
        ach_submissions.num_problems = 1
        ach_submissions.num_submissions = 1
        ach_submissions.save()
        date = ObtainedAchievement.objects.filter(user=user).values_list('obtained_date', flat=True)
        self.assertEqual(date[0], sub_1_u.creation_date)
        ObtainedAchievement.objects.all().delete()

    def test_incorrect_submit_get_achievement(self):
        """Test if an incorrect submission get an achievement"""
        client = Client()
        user = create_user('passwordmazepin', 'mazepin')
        coll = create_collection('Coleccion de cartas')
        problem = create_select_problem(coll, 'Problema')
        ach_submissions = NumSubmissionsProblemsAchievementDefinition(name={"es":'Un envio'},
                                    description={"es":'Envia una solucion para un problema'},
                                    num_problems=1, num_submissions=1)
        ach_submissions.save()
        client.login(username='mazepin', password='passwordmazepin')
        submit_select_url = reverse('judge:submit', args=[problem.pk])
        client.post(submit_select_url, {'code': 'MAL'}, follow=True)  # Validation Error, too short
        self.assertEqual(ObtainedAchievement.objects.filter(user=user).count(), 1)

    def test_return_none(self):
        """Test if an user don't solve a problem, function solved_position return None"""
        user = create_user('passwordmichu', 'michu')
        coll = create_collection('Coleccion de cartas')
        problem = create_select_problem(coll, 'Problema')
        self.assertIsNone(problem.solved_position(user))

    def test_str_method_obtained_achievement(self):
        """Test for check if __str__ return the name of the achievement"""
        achievement_definition = AchievementDefinition(name={"es":'nombre'}, description={"es":'descripcion'})
        self.assertEqual(str(achievement_definition), achievement_definition.name['es'])

    def test_download_ranking(self):
        """ Test to download excel of results """
        client = Client()
        user = create_user('2222', 'tamara')
        teacher = create_superuser('1111', 'teacher')
        group_a = create_group('1A')
        group_a.user_set.add(user)
        start = first_day_of_course(datetime(2020, 9, 1))
        end = datetime(2021, 3, 7).strftime('%Y-%m-%d')
        collection = create_collection('Coleccion 1')
        select_problem = create_select_problem(collection, 'SelectProblem ABC DEF')
        sub = Submission.objects.create(code='SELECT * FROM test where n = 1000',
                                        user=user, veredict_code=VeredictCode.WA, problem=select_problem)
        sub.save()
        Submission.objects.filter(id=sub.id).update(creation_date=datetime(2021, 3, 5))
        client.login(username=teacher.username, password='1111')
        url = reverse('judge:download_ranking', args=[collection.pk])
        response = client.get(url, {'group': group_a.id, 'start': start, 'end': end}, follow=True)

        # Teacher download ranking
        self.assertEqual(
            response.get('Content-Disposition'),
            "attachment; filename=ranking.xlsx",
        )
        self.assertEqual(
            response.get('Content-Type'),
            "application/xlsx"
        )

        with tempfile.NamedTemporaryFile(mode='w+b', buffering=-1, suffix='.xlsx') as file:
            cont = response.content
            file.write(cont)
            work = openpyxl.load_workbook(file)
            book = work.active
            self.assertIn("Colección: " + collection.name_md, book.cell(row=1, column=1).value)
            self.assertIn("1A", book.cell(row=3, column=1).value)
            self.assertIn("1", book.cell(row=5, column=1).value)
            self.assertIn(user.username, book.cell(row=5, column=2).value)
            self.assertIn("0/1 (1)", book.cell(row=5, column=3).value)
            self.assertIn("0", book.cell(row=5, column=4).value)
            self.assertIn("0", book.cell(row=5, column=5).value)

        # Date or group invalid
        response = client.get(url, {
            'group': group_a.id, 'start': start, 'end': ''}, follow=True)
        self.assertIn("Este campo es obligatorio", response.content.decode('utf-8'))
        response = client.get(url, {
            'group': group_a.id, 'start': 'eee', 'end': end}, follow=True)
        self.assertIn("Introduzca una fecha válida", response.content.decode('utf-8'))
        response = client.get(url, {
            'group': group_a.id, 'end': end}, follow=True)
        self.assertIn("Este campo es obligatorio",
                      response.content.decode('utf-8'))
        response = client.get(url,
                              {'group': '1A', 'start': start, 'end': end}, follow=True)
        self.assertIn('Introduzca un número entero', response.content.decode('utf-8'))

        # User can't download ranking
        client.logout()
        client.login(username=user.username, password='2222')
        response = client.get(url, {'group': group_a.id, 'start': start, 'end': end}, follow=True)
        self.assertIn('Forbidden', response.content.decode('utf-8'))
