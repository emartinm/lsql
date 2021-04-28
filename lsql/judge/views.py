# -*- coding: utf-8 -*-
"""
Copyright Enrique Martín <emartinm@ucm.es> 2020
Functions that process HTTP connections
"""
from datetime import timedelta, datetime

import tempfile
import os
import pathlib
from bs4 import BeautifulSoup
import openpyxl
from logzero import logger

from django.http import HttpResponseRedirect, JsonResponse, HttpResponseForbidden
from django.http.response import HttpResponse, HttpResponseNotFound
from django.urls import reverse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext_lazy as _
from django.contrib.admin.views.decorators import staff_member_required

from .exceptions import ExecutorException
from .feedback import compile_error_to_html_table, filter_expected_db
from .forms import SubmitForm, ResultForm
from .models import Collection, Problem, Submission, ObtainedAchievement, AchievementDefinition, \
    NumSubmissionsProblemsAchievementDefinition
from .oracle_driver import OracleExecutor
from .types import VeredictCode, OracleStatusCode, ProblemType
from .statistics import submissions_by_day, submission_count

# TRANSLATIONS #
# To translate the code to another language you need to create the translation file:
# 1) $ django-admin makemessages -l en
#    where "en" is the English language code.
# 2) Complete the msgstr in lsql/locale/en/LC_MESSAGES/django.po with the new translations.
# 3) $ django-admin compilemessages
#
# If the msgstr is empty, it will use the msgid instead. 2) Use these commands everytime you add a new translation
# outside the translation file. It won't remove existing translations.
# FOR WINDOWS: https://mlocati.github.io/articles/gettext-iconv-windows.html is necessary to execute the commands.


####################
# Helper functions #
####################

def get_subclass_problem(problem_id):
    """Look for problem 'pk' in the different child classes of Problem"""
    queryset = Problem.objects.filter(pk=problem_id).select_subclasses()
    return None if len(queryset) == 0 else queryset[0]


def pos(user_1, user_2):
    """ Compares if user_1 has a better ranking than user_2 """
    if user_1.resolved == user_2.resolved and user_1.score == user_2.score:
        return False
    return True


def first_day_of_course(init_course):
    """Returns the first day of the academic year"""
    first_day = datetime(init_course.year, 9, 1).strftime('%Y-%m-%d')
    if 1 <= init_course.month < 9:
        first_day = datetime(init_course.year - 1, 9, 1).strftime('%Y-%m-%d')
    return first_day


def update_user_with_scores(user_logged, user, collection, start, end):
    """Updates user object with information about submissions to problems in collection:
       - user.score (int)
       - user.collection (list of str representing "accepted submissions/all submission (first AC)") for problems in
         collection (THE SAME ORDER)
    """
    for numb in range(collection.problem_list.count()):
        num_accepted = 0
        enter = False
        attempts = 0
        problem = collection.problem_list[numb]
        user.first_AC = 0
        if user_logged.is_staff:
            subs = Submission.objects.filter(user=user,
                                             problem=problem.id,
                                             creation_date__range=[start, end + timedelta(days=1)]).order_by('pk')

        else:
            subs = Submission.objects.filter(user=user).filter(problem=problem.id).order_by('pk')
        for (submission_pos, submission) in enumerate(subs):
            if submission.veredict_code == VeredictCode.AC:
                num_accepted = num_accepted + 1
            if num_accepted == 1 and not enter:
                enter = True
                user.first_AC = submission_pos + 1
                user.score = user.score + submission_pos + 1
            attempts = attempts + 1
        update_user_attempts_problem(attempts, user, problem, num_accepted, collection, numb, enter)


def update_user_attempts_problem(attempts, user, problem, num_accepted, collection, numb, enter):
    """Updates user.collection list with problem information:  "accepted submissions/all submission (first AC)"""
    if attempts > 0 and user.first_AC == 0:
        problem.num_submissions = f"{num_accepted}/{attempts} ({attempts})"
        problem.solved = False
    else:
        problem.num_submissions = f"{num_accepted}/{attempts} ({user.first_AC})"
        problem.solved = collection.problem_list[numb].solved_by_user(user)
    if problem.solved and enter:
        user.resolved = user.resolved + 1
    user.collection.append(problem)


def check_if_get_achievement(user, veredict):
    """Check if the user get some achievement"""
    if veredict == VeredictCode.AC:
        for ach in AchievementDefinition.objects.all().select_subclasses():
            ach.check_and_save(user)
    # If the veredict != AC (correct) only can get a NumSubmissionsProblemsAchievementDefinition
    else:
        for ach in NumSubmissionsProblemsAchievementDefinition.objects.all():
            ach.check_and_save(user)


##############
#   Views    #
##############

def index(_):
    """Redirect root access to collections"""
    return HttpResponseRedirect(reverse('judge:collections'))


@login_required
def help_page(request):
    """ Shows help page """
    # FIX: handle languages
    return render(request, 'help.es.html')


@login_required
def show_result(request, collection_id):
    """show the ranking of a group (GET param) for collection_id"""
    if not Group.objects.all():
        # Show an informative message if there are not groups in the system
        return render(request, 'generic_error_message.html',
                      {'error': [_('¡Lo sentimos! No existe ningún grupo para ver resultados')]})
    position = 1
    result_form = ResultForm(request.GET)
    start = None
    end = None
    up_to_classification_date = None
    from_classification_date = None
    up_to_classification = datetime.today().strftime('%Y-%m-%d')
    collection = get_object_or_404(Collection, pk=collection_id)
    if request.user.is_staff and result_form.is_valid():
        group_id = result_form.cleaned_data['group']
        start = result_form.cleaned_data['start']
        end = result_form.cleaned_data['end']
        groups_user = Group.objects.all().order_by('name')
        up_to_classification_date = end
        from_classification_date = start
    elif request.user.is_staff and not result_form.is_valid():
        # Error 404 with all the validation errors as HTML
        return HttpResponseNotFound(str(result_form.errors))
    else:
        if result_form.is_valid():
            return HttpResponseForbidden("Forbidden")
        groups_user = request.user.groups.all().order_by('name')
        group_id = request.GET.get('group')
    if group_id is None:
        group_id = groups_user[0].id
    group0 = get_object_or_404(Group, pk=group_id)
    users = get_user_model().objects.filter(groups__name=group0.name)
    if users.filter(id=request.user.id) or request.user.is_staff:
        group0.name = group0.name
        group0.id = group_id
        groups_user = groups_user.exclude(id=group_id)
        collection.problem_list = collection.problems()
        collection.total_problem = collection.problem_list.count()
        users = users.exclude(is_staff=True)
        for user in users:
            user.collection = []
            user.resolved = 0
            user.score = 0
            user.n_achievements = ObtainedAchievement.objects.filter(user=user.pk).count()
            update_user_with_scores(request.user, user, collection, start, end)
        users = sorted(users, key=lambda x: (x.resolved, -x.score), reverse=True)
        length = len(users)
        for i in range(length):
            if i != len(users) - 1:
                if pos(users[i], users[i + 1]):
                    users[i].pos = position
                    position = position + 1
                else:
                    users[i].pos = position
            else:
                if pos(users[i], users[i - 1]):
                    users[i].pos = position
                    position = position + 1
                else:
                    users[i].pos = position
        return render(request, 'results.html', {'collection': collection, 'groups': groups_user,
                                                'users': users, 'login': request.user,
                                                'group0': group0,
                                                'to_fixed': up_to_classification,
                                                'from_date': from_classification_date,
                                                'to_date': up_to_classification_date,
                                                'end_date': end, 'start_date': start})

    return HttpResponseForbidden("Forbidden")


@login_required
def show_results(request):
    """shows the links to enter the results of each collection"""
    if not Group.objects.all():
        # Show an informative message if there are not groups in the system
        return render(request, 'generic_error_message.html',
                      {'error': [_('¡Lo sentimos! No existe ningún grupo para ver resultados')]})

    cols = Collection.objects.all().order_by('position', '-creation_date')
    groups_user = request.user.groups.all().order_by('name')
    if groups_user.count() == 0 and not request.user.is_staff:
        return render(request, 'generic_error_message.html',
                      {'error': [_('¡Lo sentimos! No tienes asignado un grupo de la asignatura.'),
                                 _('Por favor, contacta con tu profesor para te asignen un grupo de clase.')]
                       })
    if groups_user.count() == 0 and request.user.is_staff:
        groups_user = Group.objects.all().order_by('name')
    for results in cols:
        # Templates can only invoke nullary functions or access object attribute, so we store
        # the number of problems solved by the user in an attribute
        results.num_solved = results.num_solved_by_user(request.user)
    up_to_classification = datetime.today().strftime('%Y-%m-%d')
    up_to_classification_date = datetime.strptime(up_to_classification, '%Y-%m-%d')

    from_classification = first_day_of_course(datetime.today())
    from_classification_date = datetime.strptime(from_classification, '%Y-%m-%d')
    return render(request, 'result.html', {'user': request.user, 'results': cols, 'group': groups_user[0].id,
                                           'start_date': from_classification,
                                           'from_date': from_classification_date,
                                           'to_date': up_to_classification_date,
                                           'to_fixed': up_to_classification,
                                           'end_date': up_to_classification})


@login_required
def show_collections(request):
    """Show all the collections"""
    cols = Collection.objects.all().order_by('position', '-creation_date')
    for collection in cols:
        # Templates can only invoke nullary functions or access object attribute, so we store
        # the number of problems solved by the user in an attribute
        collection.num_solved = collection.num_solved_by_user(request.user)
    return render(request, 'collections.html', {'collections': cols})


@login_required
def show_collection(request, collection_id):
    """Shows a collection"""
    collection = get_object_or_404(Collection, pk=collection_id)
    # New attribute to store the list of problems and include the number of submission in each problem
    collection.problem_list = collection.problems()
    for problem in collection.problem_list:
        problem.num_submissions = problem.num_submissions_by_user(request.user)
        problem.solved = problem.solved_by_user(request.user)
    return render(request, 'collection.html', {'collection': collection})


@login_required
def show_problem(request, problem_id):
    """Shows a concrete problem"""
    # Error 404 if there is no a Problem pk
    get_object_or_404(Problem, pk=problem_id)
    # Look for problem pk in all the Problem classes
    problem = get_subclass_problem(problem_id)
    # Stores the flag in an attribute so that the template can use it
    problem.solved = problem.solved_by_user(request.user)
    # Filter the expected result to display it
    problem.show_added, problem.show_modified, problem.show_removed = filter_expected_db(problem.expected_result[0],
                                                                                         problem.initial_db[0])
    return render(request, problem.template(), {'problem': problem})


@login_required
def show_submissions(request):
    """Shows all the submissions of the current user"""

    try:
        pk_problem = request.GET.get('problem_id')
        id_user = request.GET.get('user_id')
        if pk_problem is not None or id_user is not None:
            if int(id_user) == request.user.id and not request.user.is_staff:
                start = request.GET.get('start')
                end = request.GET.get('end')
                if start is not None or end is not None:
                    return HttpResponseForbidden("Forbidden")
                problem = get_object_or_404(Problem, pk=pk_problem)
                subs = Submission.objects.filter(user=request.user).filter(problem=problem.id).order_by('-pk')
            elif request.user.is_staff:
                start = request.GET.get('start')
                end = request.GET.get('end')
                starts = datetime.strptime(start, '%Y-%m-%d')
                ends = datetime.strptime(end, '%Y-%m-%d')
                problem = get_object_or_404(Problem, pk=pk_problem)
                user = get_user_model().objects.filter(id=id_user)
                subs = Submission.objects.filter(user=user.get()) \
                    .filter(problem=problem.id, creation_date__range=[starts, ends + timedelta(days=1)]).order_by('-pk')
            else:
                return HttpResponseForbidden("Forbidden")

        else:
            subs = Submission.objects.filter(user=request.user).order_by('-pk')
        for submission in subs:
            submission.veredict_pretty = VeredictCode(submission.veredict_code).html_short_name()
        return render(request, 'submissions.html', {'submissions': subs})
    except ValueError:
        return HttpResponseNotFound(_("El identificador no tiene el formato correcto"))


@login_required
def show_achievements(request, user_id):
    """View for show the achievements"""
    this_user = get_user_model().objects.get(pk=user_id)
    achievements_locked = []
    achievements_unlocked = ObtainedAchievement.objects.filter(user=this_user)
    achievements_definitions_unlocked = ObtainedAchievement.objects.filter(user=this_user).\
        values_list('achievement_definition', flat=True)
    all_achievements_definitions = AchievementDefinition.objects.values_list('pk', flat=True)
    achievements_locked_pk = all_achievements_definitions.difference(achievements_definitions_unlocked)
    for identifier in achievements_locked_pk:
        ach = AchievementDefinition.objects.get(pk=identifier)
        achievements_locked.append(ach)
    return render(request, 'achievements.html', {'locked': achievements_locked,
                                                 'unlocked': achievements_unlocked,
                                                 'username': this_user.username})


@login_required
def show_submission(request, submission_id):
    """Shows a submission of the current user"""
    submission = get_object_or_404(Submission, pk=submission_id)
    if submission.user != request.user and not request.user.is_staff:
        return HttpResponseForbidden("Forbidden")
    submission.veredict_pretty = VeredictCode(submission.veredict_code).html_short_name()
    return render(request, 'submission.html', {'submission': submission})


@login_required
def download_submission(request, submission_id):
    """ Return a script with te code of submission """
    submission = get_object_or_404(Submission, pk=submission_id)
    if submission.user != request.user and not request.user.is_staff:
        return HttpResponseForbidden("Forbidden")
    response = HttpResponse()
    response['Content-Type'] = 'application/sql'
    response['Content-Disposition'] = "attachment; filename=code.sql"
    response.write(submission.code)
    return response


@login_required
def download(_, problem_id):
    """Returns a script with the creation and insertion of the problem"""
    get_object_or_404(Problem, pk=problem_id)
    # Look for problem pk in all the Problem classes
    problem = get_subclass_problem(problem_id)
    response = HttpResponse()
    response['Content-Type'] = 'application/sql'
    response['Content-Disposition'] = "attachment; filename=create_insert.sql"
    response.write(problem.create_sql + '\n\n' + problem.insert_sql)
    return response


@login_required
# pylint does not understand the dynamic attributes in VeredictCode (TextChoices), so we need to disable
# no-member warning in this specific function
# pylint: disable=no-member
def submit(request, problem_id):
    """Process a user submission"""
    # Error 404 if there is no Problem 'pk'
    general_problem = get_object_or_404(Problem, pk=problem_id)
    problem = get_subclass_problem(problem_id)
    submit_form = SubmitForm(request.POST)
    data = {'veredict': VeredictCode.IE, 'title': VeredictCode.IE.label,
            'message': VeredictCode.IE.message(), 'feedback': ''}
    code = ''
    if submit_form.is_valid():
        try:
            # AC or WA
            code = submit_form.cleaned_data['code']
            data['veredict'], data['feedback'] = problem.judge(code, OracleExecutor.get())
            data['title'] = data['veredict'].label
            data['message'] = data['veredict'].message()
        except ExecutorException as excp:
            # Exceptions when judging: RE, TLE, VE or IE
            if excp.error_code == OracleStatusCode.EXECUTE_USER_CODE:
                data = {
                    'veredict': VeredictCode.RE,
                    'title': VeredictCode.RE.label,
                    'message': VeredictCode.RE.message(),
                    'feedback': (f'{excp.statement} --> {excp.message}'
                                 if problem.problem_type() == ProblemType.FUNCTION else excp.message),
                    'position': excp.position,
                    'position_msg': _('Posición: línea {row}, columna {col}').format(row=excp.position[0]+1,
                                                                                     col=excp.position[1]+1)
                }
            elif excp.error_code == OracleStatusCode.TLE_USER_CODE:
                data = {'veredict': VeredictCode.TLE, 'title': VeredictCode.TLE.label,
                        'message': VeredictCode.TLE.message(), 'feedback': ''}
            elif excp.error_code == OracleStatusCode.NUMBER_STATEMENTS:
                data = {'veredict': VeredictCode.VE, 'title': VeredictCode.VE.label,
                        'message': VeredictCode.VE.message(problem), 'feedback': excp.message}
            elif excp.error_code == OracleStatusCode.COMPILATION_ERROR:
                data = {'veredict': VeredictCode.WA, 'title': VeredictCode.WA.label,
                        'message': VeredictCode.WA.message(), 'feedback': compile_error_to_html_table(excp.message)}
    else:
        data = {'veredict': VeredictCode.VE, 'title': VeredictCode.VE.label,
                'message': VeredictCode.VE.message(), 'feedback': ''}

    submission = Submission(code=code, veredict_code=data['veredict'], veredict_message=data['message'],
                            user=request.user, problem=general_problem)
    submission.save()
    # If verdict is correct look for an achievement to complete if it's possible
    check_if_get_achievement(request.user, data['veredict'])
    logger.debug('Stored submission %s', submission)
    return JsonResponse(data)


@login_required
def password_change_done(request):
    """Password change confirmation"""
    return render(request, 'password_change_done.html')


def test_error_500(request):
    """Generates a server internal error, only for testing error reporting in deployment"""
    if request.user and request.user.is_staff:
        elems = list()
        return HttpResponse(elems[55])  # list index out of range, error 500
    return HttpResponseNotFound()


@login_required
def download_ranking(request, collection_id):
    """Download excel with the results of submissions"""
    result_form = ResultForm(request.GET)

    if request.user.is_staff and result_form.is_valid():
        html = show_result(request, collection_id).content.decode('utf-8')
        soup = BeautifulSoup(html, 'html.parser')
        col = 1
        row = 1
        work = openpyxl.Workbook()
        book = work.active

        # Takes collection and date
        data = soup.find_all("h1")
        for i in data:
            book.cell(row=row, column=col, value=i.string.strip())
            row += 1

        # Takes group
        option = soup.find(id='clase').find("option")
        book.cell(row=row, column=col, value=option.string.strip())
        row += 1

        # Takes the first column of table (Pos, User, Exercises, Score, Solved)
        ths = soup.find_all("th")
        for i in ths:
            if i.string is not None:
                book.cell(row=row, column=col, value=i.string.strip())
                col += 1
            exercises = i.find_all("a")
            for j in exercises:
                if j.string is not None:
                    book.cell(row=row, column=col, value=j['title'].strip())
                    col += 1
        col = 1
        # Takes the information for each student
        trs = soup.find_all("tr")
        for i in trs:
            tds = i.find_all("td")
            # Information of a student (Pos, User, Exercises, Score, Solved)
            for j in tds:
                name = j.find('span', class_='ranking-username')
                if name is not None:
                    book.cell(row=row, column=col, value=name.string.strip())
                    col += 1
                if j.string is not None:
                    book.cell(row=row, column=col, value=j.string.strip())
                    col += 1
                num_submissions = j.find_all("a")
                for k in num_submissions:
                    if k.string is not None:
                        book.cell(row=row, column=col, value=k.string.strip())
                        col += 1
            col = 1
            row += 1

        # create a temporary file to save the workbook with the results
        temp = tempfile.NamedTemporaryFile(mode='w+b', buffering=-1, suffix='.xlsx')
        file = pathlib.Path(temp.name)
        name = file.name
        work.save(name)
        response = HttpResponse(open(name, 'rb').read())
        response['Content-Type'] = 'application/xlsx'
        response['Content-Disposition'] = "attachment; filename=ranking.xlsx"
        temp.close()
        os.remove(name)
        return response

    if request.user.is_staff and not result_form.is_valid():
        return HttpResponseNotFound(str(result_form.errors))
    return HttpResponseForbidden("Forbidden")


@staff_member_required
def statistics_submissions(request):
    """ Shows statistics page containing charts and other summarized information """
    all_submissions_count = submissions_by_day()
    start = all_submissions_count[0][0] if all_submissions_count else None
    end = all_submissions_count[-1][0] if all_submissions_count else None
    ac_submissions = submissions_by_day(start, end, verdict_code=VeredictCode.AC)
    wa_submissions = submissions_by_day(start, end, verdict_code=VeredictCode.WA)
    re_submissions = submissions_by_day(start, end, verdict_code=VeredictCode.RE)
    sub_count = submission_count()
    return render(request, 'statistics_submissions.html',
                  {'all_submissions_count': all_submissions_count,
                   'ac_submissions_count': ac_submissions,
                   'wa_submissions_count': wa_submissions,
                   're_submissions_count': re_submissions,
                   'submission_count': sub_count})
